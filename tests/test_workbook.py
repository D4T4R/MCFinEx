import pytest
from openpyxl import Workbook, load_workbook

from mcfinex.db.store import Store
from mcfinex.export import workbook as wb
from mcfinex.export.workbook import WorkbookError, populate, tickers_in


@pytest.fixture
def template(tmp_path):
    """A stand-in for SSP_Working_merged.xlsx: 3 header rows, formulas intact."""
    book = Workbook()
    sheet = book.active
    sheet.title = wb.DATA_SHEET
    sheet.cell(3, wb.TICKER_COLUMN).value = "TICKER"
    # One existing company, plus a formula that must survive the round trip.
    sheet.cell(4, wb.TICKER_COLUMN).value = "ACME"
    sheet.cell(4, 10).value = "=H4/I4"
    path = tmp_path / "template.xlsx"
    book.save(path)
    return path


@pytest.fixture
def store(tmp_path):
    with Store(tmp_path / "test.db") as s:
        s.create_schema()
        s.upsert_company("ACME", {
            "name": "Acme Ltd", "current_price": 100.0, "market_cap": 5000.0,
            "outstanding_shares": 50.0, "last_updated": "2026-01-01",
            "last_updated_quarter": "2026-4", "industry": "Widgets",
        })
        s.replace_financials("ACME", [
            ("2022-03-31", "profit-loss", "Operating Profit", 10.0),
            ("2023-03-31", "profit-loss", "Operating Profit", 20.0),
            ("2024-03-31", "profit-loss", "Operating Profit", 30.0),
            ("2023-03-31", "profit-loss", "EPS in Rs", 4.0),
            ("2024-03-31", "profit-loss", "EPS in Rs", 5.0),
            ("2024-03-31", "balance-sheet", "Reserves", 900.0),
            ("2024-03-31", "balance-sheet", "Borrowings", 250.0),
            ("2024-06-30", "quarters", "EPS in Rs", 1.0),
            ("2024-09-30", "quarters", "EPS in Rs", 2.0),
            ("2024-12-31", "quarters", "EPS in Rs", 3.0),
            ("2025-03-31", "quarters", "EPS in Rs", 4.0),
        ])
        s.replace_valuations("ACME", "derived", {"ebit": 77.0, "free_cash_flow": -5.0})
        yield s


def read(path, row, column):
    return load_workbook(path)[wb.DATA_SHEET].cell(row, column).value


class TestPopulate:
    def test_updates_an_existing_row_in_place(self, store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        _, updated, appended = populate(store, template, out)
        assert (updated, appended) == (1, 0)
        assert read(out, 4, wb.COMPANY_NAME) == "Acme Ltd"

    def test_leaves_the_template_untouched(self, store, template, tmp_path):
        populate(store, template, tmp_path / "out.xlsx")
        assert read(template, 4, wb.COMPANY_NAME) is None

    def test_in_place_writes_to_the_template(self, store, template, tmp_path):
        populate(store, template, tmp_path / "unused.xlsx", in_place=True)
        assert read(template, 4, wb.COMPANY_NAME) == "Acme Ltd"

    def test_preserves_existing_formulas(self, store, template, tmp_path):
        # The workbook is the model; the scraper must not overwrite its maths.
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        assert read(out, 4, 10) == "=H4/I4"

    def test_appends_an_unknown_ticker_to_the_first_blank_row(self, store, template, tmp_path):
        store.upsert_company("NEWCO", {"name": "New Co", "last_updated": "2026-01-01"})
        out = tmp_path / "out.xlsx"
        _, updated, appended = populate(store, template, out)
        assert appended == 1
        assert read(out, 5, wb.TICKER_COLUMN) == "NEWCO"

    def test_missing_template_is_reported(self, store, tmp_path):
        with pytest.raises(WorkbookError, match="template not found"):
            populate(store, tmp_path / "nope.xlsx", tmp_path / "out.xlsx")

    def test_sheet_must_exist(self, store, tmp_path):
        book = Workbook()
        book.active.title = "Something Else"
        bad = tmp_path / "bad.xlsx"
        book.save(bad)
        with pytest.raises(WorkbookError, match="no 'Data' sheet"):
            populate(store, bad, tmp_path / "out.xlsx")


class TestFormulaRowLimit:
    """Appending past the template's formula range produces dead rows."""

    def _template_with_formulas(self, tmp_path, last_row):
        book = Workbook()
        sheet = book.active
        sheet.title = wb.DATA_SHEET
        for row in range(4, last_row + 1):
            sheet.cell(row, wb.TARGET_PRICE).value = f"=BP{row}/BS{row}"
        path = tmp_path / "limited.xlsx"
        book.save(path)
        return path

    def _store_with(self, tmp_path, tickers):
        s = Store(tmp_path / "many.db")
        s.create_schema()
        for t in tickers:
            s.upsert_company(t, {"name": t.title(), "last_updated": "2026-01-01"})
        return s

    def test_companies_beyond_the_formula_range_are_skipped(self, tmp_path):
        template = self._template_with_formulas(tmp_path, 6)  # rows 4,5,6
        store = self._store_with(tmp_path, ["AAA", "BBB", "CCC", "DDD", "EEE"])
        _, updated, appended = populate(store, template, tmp_path / "out.xlsx")
        store.close()
        assert (updated, appended) == (0, 3)  # only the formula rows are filled

    def test_a_sheet_with_no_formulas_is_not_capped(self, tmp_path):
        book = Workbook()
        book.active.title = wb.DATA_SHEET
        bare = tmp_path / "bare.xlsx"
        book.save(bare)
        store = self._store_with(tmp_path, ["AAA", "BBB", "CCC"])
        _, _, appended = populate(store, bare, tmp_path / "out.xlsx")
        store.close()
        assert appended == 3

    def test_skipped_rows_are_left_empty_not_half_written(self, tmp_path):
        template = self._template_with_formulas(tmp_path, 5)
        store = self._store_with(tmp_path, ["AAA", "BBB", "CCC"])
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        store.close()
        assert read(out, 6, wb.TICKER_COLUMN) is None
        assert read(out, 6, wb.COMPANY_NAME) is None


class TestTickersIn:
    def test_reads_the_tracked_tickers(self, template):
        assert tickers_in(template) == ["ACME"]

    def test_skips_header_rows_and_blanks(self, tmp_path):
        book = Workbook()
        sheet = book.active
        sheet.title = wb.DATA_SHEET
        sheet.cell(3, wb.TICKER_COLUMN).value = "TICKER"   # header, must not appear
        sheet.cell(4, wb.TICKER_COLUMN).value = " acme "   # trimmed and upper-cased
        sheet.cell(5, wb.TICKER_COLUMN).value = "   "      # blank
        sheet.cell(6, wb.TICKER_COLUMN).value = "BETA"
        path = tmp_path / "t.xlsx"
        book.save(path)
        assert tickers_in(path) == ["ACME", "BETA"]

    def test_deduplicates(self, tmp_path):
        book = Workbook()
        sheet = book.active
        sheet.title = wb.DATA_SHEET
        for row, value in enumerate(["ACME", "BETA", "ACME"], start=4):
            sheet.cell(row, wb.TICKER_COLUMN).value = value
        path = tmp_path / "t.xlsx"
        book.save(path)
        assert tickers_in(path) == ["ACME", "BETA"]

    def test_missing_template_is_reported(self, tmp_path):
        with pytest.raises(WorkbookError, match="template not found"):
            tickers_in(tmp_path / "nope.xlsx")


class TestSeriesLayout:
    def test_series_are_written_newest_first(self, store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        # CD is Y5, the latest year, because CP (current P/E) divides by it.
        assert read(out, 4, wb.EPS_YEARLY_FIRST) == 5.0
        assert read(out, 4, wb.EPS_YEARLY_FIRST + 1) == 4.0
        assert read(out, 4, wb.EBITDA_FIRST) == 30.0
        assert read(out, 4, wb.EBITDA_FIRST + 2) == 10.0

    def test_ttm_eps_sums_the_four_newest_quarters(self, store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        assert read(out, 4, wb.EPS_TTM) == 10.0  # 4 + 3 + 2 + 1

    def test_latest_balance_sheet_values_are_used(self, store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        assert read(out, 4, wb.RESERVES) == 900.0
        assert read(out, 4, wb.LONG_TERM_BORROWINGS) == 250.0

    def test_debt_column_carries_borrowings(self, store, template, tmp_path):
        # N is headed "CURRENT LIABILITY" but holds debt, so that
        # P=(M+N)/O is a real debt-to-equity ratio. See workbook.DEBT.
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        assert read(out, 4, wb.DEBT) == 250.0

    def test_current_assets_and_liabilities_are_never_written(self, store, template, tmp_path):
        # Screener has no current/non-current split; writing a proxy here would
        # silently corrupt the current ratio.
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        assert read(out, 4, wb.CURRENT_ASSETS) is None
        assert read(out, 4, wb.CURRENT_LIABILITY_2) is None

    def test_derived_values_land_in_their_columns(self, store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        assert read(out, 4, wb.EBIT) == 77.0
        assert read(out, 4, wb.FREE_CASHFLOW) == -5.0

    def test_unsourceable_columns_are_left_alone(self, tmp_path, store):
        # Industry P/E and promoter pledge have no screener source, so whatever
        # the user already has must survive.
        book = Workbook()
        sheet = book.active
        sheet.title = wb.DATA_SHEET
        sheet.cell(4, wb.TICKER_COLUMN).value = "ACME"
        sheet.cell(4, wb.INDUSTRY_PE).value = 29.47
        sheet.cell(4, wb.PROMOTER_PLEDGE).value = 3.5
        path = tmp_path / "keep.xlsx"
        book.save(path)
        out = tmp_path / "out.xlsx"
        populate(store, path, out)
        assert read(out, 4, wb.INDUSTRY_PE) == 29.47
        assert read(out, 4, wb.PROMOTER_PLEDGE) == 3.5

    def test_owned_column_with_no_data_is_cleared_not_left_stale(self, tmp_path, store):
        # A row must be wholly fresh or blank; never fresh beside a 2022 value.
        book = Workbook()
        sheet = book.active
        sheet.title = wb.DATA_SHEET
        sheet.cell(4, wb.TICKER_COLUMN).value = "ACME"
        sheet.cell(4, wb.INVENTORY_TURNOVER).value = -8888888
        path = tmp_path / "stale.xlsx"
        book.save(path)
        out = tmp_path / "out.xlsx"
        populate(store, path, out)
        assert read(out, 4, wb.INVENTORY_TURNOVER) is None

    def test_orphaned_ev_columns_are_cleared(self, tmp_path, store):
        book = Workbook()
        sheet = book.active
        sheet.title = wb.DATA_SHEET
        sheet.cell(4, wb.TICKER_COLUMN).value = "ACME"
        for column in (47, 51, 53, 56):
            sheet.cell(4, column).value = -8888888
        path = tmp_path / "orphan.xlsx"
        book.save(path)
        out = tmp_path / "out.xlsx"
        populate(store, path, out)
        assert [read(out, 4, c) for c in (47, 51, 53, 56)] == [None, None, None, None]

    def test_az_is_not_swept_up_by_the_orphan_clearing(self):
        # AZ sits between the two orphaned spans and must survive, because
        # BP=AZ*BO still consumes it.
        spans = (wb.ORPHANED_EV_RANGE, wb.ORPHANED_MULTIPLE_RANGE)
        assert not any(start <= wb.EV_EBITDA_MULTIPLE <= end for start, end in spans)

    def test_az_receives_the_derived_multiple(self, tmp_path, template):
        with Store(tmp_path / "m.db") as s:
            s.create_schema()
            s.upsert_company("ACME", {"name": "Acme", "last_updated": "2026-01-01"})
            s.replace_valuations("ACME", "derived", {"ev_ebitda_multiple": 12.5})
            out = tmp_path / "out.xlsx"
            populate(s, template, out)
        assert read(out, 4, wb.EV_EBITDA_MULTIPLE) == 12.5

    def test_short_series_clears_the_trailing_cells(self, store, template, tmp_path):
        # Only three EBITDA periods exist, so BH and BI must be blanked rather
        # than left holding the template's =AY/BD formula over sentinels.
        out = tmp_path / "out.xlsx"
        populate(store, template, out)
        assert read(out, 4, wb.EBITDA_FIRST + 3) is None
        assert read(out, 4, wb.EBITDA_FIRST + 4) is None

    def test_stale_formula_in_a_series_cell_is_overwritten(self, tmp_path, store):
        book = Workbook()
        sheet = book.active
        sheet.title = wb.DATA_SHEET
        sheet.cell(4, wb.TICKER_COLUMN).value = "ACME"
        sheet.cell(4, wb.EBITDA_FIRST + 4).value = "=AY4/BD4"  # Java-era formula
        path = tmp_path / "stale.xlsx"
        book.save(path)
        out = tmp_path / "out.xlsx"
        populate(store, path, out)
        assert read(out, 4, wb.EBITDA_FIRST + 4) is None


class TestBankLayout:
    """Screener labels banks differently; those rows must still populate."""

    @pytest.fixture
    def bank_store(self, tmp_path):
        with Store(tmp_path / "bank.db") as s:
            s.create_schema()
            s.upsert_company("BANKX", {"name": "Bank X", "last_updated": "2026-01-01"})
            s.replace_financials("BANKX", [
                ("2025-03-31", "balance-sheet", "Reserves", 200.0),
                ("2025-03-31", "balance-sheet", "Equity Capital", 10.0),
                ("2025-03-31", "balance-sheet", "Other Liabilities", 100.0),
                ("2025-03-31", "balance-sheet", "Deposits", 1300.0),
                ("2025-03-31", "balance-sheet", "Borrowing", 230.0),
                ("2025-03-31", "balance-sheet", "Total Liabilities", 1840.0),
                ("2025-03-31", "profit-loss", "Financing Profit", 45.0),
            ])
            yield s

    def test_singular_borrowing_is_matched(self, bank_store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        populate(bank_store, template, out, tickers=["BANKX"])
        assert read(out, 5, wb.DEBT) == 230.0
        assert read(out, 5, wb.LONG_TERM_BORROWINGS) == 230.0

    def test_deposits_are_folded_into_other_liabilities(self, bank_store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        populate(bank_store, template, out, tickers=["BANKX"])
        assert read(out, 5, wb.OTHER_LIABILITY) == 1400.0

    def test_balance_sheet_reconciles(self, bank_store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        populate(bank_store, template, out, tickers=["BANKX"])
        total = sum(read(out, 5, c) for c in
                    (wb.RESERVES, wb.EQUITY_CAPITAL, wb.OTHER_LIABILITY, wb.DEBT))
        assert total == read(out, 5, wb.TOTAL_LIABILITY)

    def test_financing_profit_fills_the_ebitda_series(self, bank_store, template, tmp_path):
        out = tmp_path / "out.xlsx"
        populate(bank_store, template, out, tickers=["BANKX"])
        assert read(out, 5, wb.EBITDA_FIRST) == 45.0

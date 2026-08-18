"""Populate the SSP working workbook.

The workbook *is* the model. Columns such as ``BE=AU/AZ``,
``CI=IFERROR((CD-CE)/CE*SIGN(CE),0)`` and ``CS=CO+(CO*CR%)`` already implement
the EV/EBITDA and EPS valuations, and the Results sheet is nothing but
references back into Data. So this writes raw scraped inputs into the input
cells and leaves every formula alone -- Excel recalculates on open.

Layout of the Data sheet:

* rows 1-3 are headers (row 1 carries the 0-based POI indices the Java used)
* row 4 onwards is one company per row, keyed by the ticker in column B
* columns A-AT are the SSP fundamental screen
* AU-CC is EV/EBITDA, CD-CT yearly EPS, CU-DL quarterly EPS

Series columns run **newest first**: CD is Y5 and ``CP = CO/CD`` is the current
P/E, so Y5/Q5 are the latest period.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from .. import labels
from ..db.store import Store

DATA_SHEET = "Data"
FIRST_DATA_ROW = 4
TICKER_COLUMN = 2  # column B

# --- single-value cells, by 1-based column index -------------------------
COMPANY_NAME = 3          # C
PROMOTER_HOLDING = 4      # D
PROMOTER_PLEDGE = 5       # E
RESERVES = 8              # H
EQUITY_CAPITAL = 9        # I
OTHER_LIABILITY = 13      # M
# N is headed "CURRENT LIABILITY", but screener does not split current from
# non-current -- its "Other Liabilities" already includes both, and lands in M.
# N therefore carries Borrowings so that P=(M+N)/O is a correct debt-to-equity
# ratio, which is what that column group is actually titled. The header text is
# left stale on purpose rather than editing the user's template.
DEBT = 14                 # N
CURRENT_ASSETS = 18       # R, no screener source; left untouched
CURRENT_LIABILITY_2 = 19  # S, no screener source; feeds T=R/S and X=W-S
EBIT = 22                 # V
TOTAL_LIABILITY = 23      # W
INVENTORY_TURNOVER = 27   # AA
OPERATING_CASHFLOW = 30   # AD
INVESTING_CASHFLOW = 31   # AE
FINANCING_CASHFLOW = 32   # AF
FREE_CASHFLOW = 34        # AH
MARKET_PRICE = 36         # AJ
EPS_TTM = 37              # AK
INDUSTRY_PE = 39          # AM
BOOK_VALUE = 42           # AP
DIVIDEND_YIELD = 45       # AS
# AU-AY (enterprise value) and BA-BD (EV/EBITDA 2-5) fed BE=AU/AZ under
# MoneyControl. Screener publishes no enterprise-value history, so EBITDA goes
# straight into BE-BI and nothing points at this range any more. It is cleared
# rather than abandoned: it is full of -8888888 sentinels that would otherwise
# sit beside current figures looking like data.
ORPHANED_EV_RANGE = (47, 51)        # AU-AY
ORPHANED_MULTIPLE_RANGE = (53, 56)  # BA-BD
EV_EBITDA_MULTIPLE = 52   # AZ, the EV/EBITDA 1 that BP=AZ*BO consumes
EBITDA_FIRST = 57         # BE..BI
LONG_TERM_BORROWINGS = 69  # BQ
SHARES_OUTSTANDING = 71   # BS
EPS_YEARLY_FIRST = 82     # CD..CH  (Y5..Y1)
EPS_QUARTERLY_FIRST = 99  # CU..CY  (Q5..Q1)
MARKET_CAPITAL = 117      # DM
SECTOR = 118              # DN
LAST_UPDATED_QUARTER = 119  # DO
LAST_CHECKED_ON = 120     # DP
SYMBOL = 121              # DQ

SERIES_LENGTH = 5


class WorkbookError(RuntimeError):
    pass


def populate(store: Store, template: str | Path, output: str | Path,
             *, tickers: list[str] | None = None, in_place: bool = False) -> tuple[Path, int, int]:
    """Write scraped data into a copy of ``template``.

    Returns ``(path, updated, appended)``. Unless ``in_place`` is set the
    template is copied to ``output`` first, so the source workbook is never the
    thing being mutated.
    """
    template, output = Path(template), Path(output)
    if not template.exists():
        raise WorkbookError(f"template not found: {template}")

    target = template if in_place else output
    if not in_place:
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(template, target)

    book = load_workbook(target)
    if DATA_SHEET not in book.sheetnames:
        raise WorkbookError(f"{template} has no '{DATA_SHEET}' sheet")
    sheet = book[DATA_SHEET]
    _force_recalculation(book)

    index = _ticker_index(sheet)
    next_free = _first_blank_row(sheet)

    updated = appended = 0
    for ticker in (tickers if tickers is not None else store.scraped_tickers()):
        row = index.get(ticker)
        if row is None:
            row = next_free
            next_free += 1
            sheet.cell(row, TICKER_COLUMN).value = ticker
            appended += 1
        else:
            updated += 1
        _write_company(sheet, row, store, ticker)

    book.save(target)
    return target, updated, appended


def _force_recalculation(book) -> None:
    """Make the workbook recalculate when it is opened.

    Two things conspire to leave every formula reading zero. The template ships
    with ``calcMode="manual"``, so the application will not recalculate on its
    own; and openpyxl writes formulas without their cached results, so there is
    no previous value to fall back on. The whole EV/EBITDA chain -- BN, BO, BP,
    BR and the BT-CC target prices that depend on them -- therefore displays 0
    even though the inputs are correct.

    Setting automatic calculation as well as ``fullCalcOnLoad`` covers both
    Excel and LibreOffice, which ignores the latter on its own.
    """
    book.calculation.calcMode = "auto"
    book.calculation.fullCalcOnLoad = True


def tickers_in(template: str | Path) -> list[str]:
    """Every ticker already listed in the workbook, in sheet order.

    Lets a scrape target exactly the companies being tracked rather than the
    whole NSE universe.
    """
    template = Path(template)
    if not template.exists():
        raise WorkbookError(f"template not found: {template}")
    book = load_workbook(template, read_only=True)
    if DATA_SHEET not in book.sheetnames:
        raise WorkbookError(f"{template} has no '{DATA_SHEET}' sheet")

    found: list[str] = []
    seen: set[str] = set()
    rows = book[DATA_SHEET].iter_rows(
        min_row=FIRST_DATA_ROW, min_col=TICKER_COLUMN, max_col=TICKER_COLUMN, values_only=True
    )
    for (value,) in rows:
        if not isinstance(value, str) or not value.strip():
            continue
        ticker = value.strip().upper()
        if ticker not in seen:
            seen.add(ticker)
            found.append(ticker)
    book.close()
    return found


def _ticker_index(sheet) -> dict[str, int]:
    """Map each ticker already in column B to its row."""
    index: dict[str, int] = {}
    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        value = sheet.cell(row, TICKER_COLUMN).value
        if isinstance(value, str) and value.strip():
            index[value.strip().upper()] = row
    return index


def _first_blank_row(sheet) -> int:
    for row in range(FIRST_DATA_ROW, sheet.max_row + 2):
        value = sheet.cell(row, TICKER_COLUMN).value
        if value is None or (isinstance(value, str) and not value.strip()):
            return row
    return sheet.max_row + 1


def _write_company(sheet, row: int, store: Store, ticker: str) -> None:
    company = store.company(ticker)
    if company is None:
        return
    derived = store.valuation_fields(ticker, "derived")

    def put(column: int, value) -> None:
        """Write a value, blanking the cell when there is nothing to write.

        Clearing matters as much as writing. These are columns this tool owns,
        so a row must end up wholly fresh or blank -- never fresh in one cell
        and 2022 in the next. The template still carries Java-era sentinels
        (``-8888888``, ``-5555555``) in exactly these positions, and skipping a
        ``None`` would leave one sitting beside current data.

        Columns with no screener source (promoter pledge, industry P/E) are
        never passed here, so whatever they already hold survives.
        """
        sheet.cell(row, column).value = value

    bs, cf, pl = "balance-sheet", "cash-flow", "profit-loss"

    put(COMPANY_NAME, company["name"])
    put(SYMBOL, ticker)
    put(MARKET_PRICE, company["current_price"])
    put(MARKET_CAPITAL, company["market_cap"])
    put(BOOK_VALUE, company["book_value"])
    put(DIVIDEND_YIELD, company["dividend_yield"])
    # INDUSTRY_PE is deliberately not written: screener loads the peers table
    # separately, so there is no value for it and clearing would destroy
    # whatever the user already has.
    put(SECTOR, company["industry"] or company["sector"])
    put(LAST_UPDATED_QUARTER, company["last_updated_quarter"])
    put(LAST_CHECKED_ON, company["last_updated"])
    put(SHARES_OUTSTANDING, company["outstanding_shares"])

    put(PROMOTER_HOLDING, _latest(store, ticker, "shareholding", *labels.PROMOTERS))
    put(RESERVES, _latest(store, ticker, bs, *labels.RESERVES))
    put(EQUITY_CAPITAL, _latest(store, ticker, bs, *labels.EQUITY_CAPITAL))
    # Deposits are a bank's main liability and screener lists them separately,
    # so they are folded in here; without them the sheet fails to reconcile by
    # the entire deposit base. Summed from stored facts rather than from the
    # derived record, so a mapping fix does not require re-scraping.
    put(OTHER_LIABILITY, _sum(
        _latest(store, ticker, bs, *labels.OTHER_LIABILITIES),
        _latest(store, ticker, bs, *labels.DEPOSITS),
    ))
    put(TOTAL_LIABILITY, _latest(store, ticker, bs, *labels.TOTAL_LIABILITIES))
    borrowings = _latest(store, ticker, bs, *labels.BORROWINGS)
    put(DEBT, borrowings)
    put(LONG_TERM_BORROWINGS, borrowings)

    put(OPERATING_CASHFLOW, _latest(store, ticker, cf, *labels.CASH_FROM_OPERATING))
    put(INVESTING_CASHFLOW, _latest(store, ticker, cf, *labels.CASH_FROM_INVESTING))
    put(FINANCING_CASHFLOW, _latest(store, ticker, cf, *labels.CASH_FROM_FINANCING))
    put(FREE_CASHFLOW, derived.get("free_cash_flow"))

    put(EBIT, derived.get("ebit"))
    put(INVENTORY_TURNOVER, derived.get("inventory_turnover"))
    put(EV_EBITDA_MULTIPLE, derived.get("ev_ebitda_multiple"))

    for start, end in (ORPHANED_EV_RANGE, ORPHANED_MULTIPLE_RANGE):
        for column in range(start, end + 1):
            put(column, None)

    _write_series(sheet, row, EBITDA_FIRST,
                  store.series(ticker, pl, *labels.OPERATING_PROFIT, limit=SERIES_LENGTH))
    _write_series(sheet, row, EPS_YEARLY_FIRST,
                  store.series(ticker, pl, *labels.EPS, limit=SERIES_LENGTH))
    quarterly = store.series(ticker, "quarters", *labels.EPS, limit=SERIES_LENGTH)
    _write_series(sheet, row, EPS_QUARTERLY_FIRST, quarterly)
    # AK is the trailing-twelve-month EPS the P/E columns divide by.
    if len(quarterly) >= 4:
        put(EPS_TTM, sum(quarterly[:4]))


def _latest(store: Store, ticker: str, statement: str, *label_aliases: str) -> float | None:
    values = store.series(ticker, statement, *label_aliases, limit=1)
    return values[0] if values else None


def _sum(*values: float | None) -> float | None:
    """Total the values that are present, or ``None`` if none are."""
    present = [v for v in values if v is not None]
    return sum(present) if present else None


def _write_series(sheet, row: int, first_column: int, values: list[float]) -> None:
    """Write five cells left to right, newest first, blanking any shortfall.

    Cells beyond the available history are cleared rather than skipped. The
    template ships these columns as formulas over Java-era sentinels
    (``BI = AY/BD`` where both are ``-8888888``), so leaving one behind would
    mix a fabricated number into a row of otherwise fresh data.
    """
    for offset in range(SERIES_LENGTH):
        value = values[offset] if offset < len(values) else None
        sheet.cell(row, first_column + offset).value = value
